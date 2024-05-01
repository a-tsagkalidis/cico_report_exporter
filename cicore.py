import xlrd
import argparse
from tqdm import tqdm
from openpyxl import Workbook
from datetime import datetime


def get_timestamps_by_person(file_path, sheet_name):
    try:
        # Open the Excel file
        workbook = xlrd.open_workbook(file_path)
        
        # Access the specific sheet
        sheet = workbook.sheet_by_name(sheet_name)
        
        # Initialize a dictionary to store timestamps by person
        timestamps_by_person = {}
        
        # Iterate over rows
        for row_idx in tqdm(range(1, sheet.nrows)):
            name = sheet.cell_value(row_idx, 3)  # Assuming name is in column D
            timestamp_str = sheet.cell_value(row_idx, 5)  # Assuming timestamp is in column F
            
            # Correcting the timestamp format to include the day of the week
            timestamp = datetime.strptime(timestamp_str, "%Y-%m-%d %H:%M:%S %A")
            
            # Extract year, month, and day
            year = str(timestamp.year)
            month = timestamp.strftime("%B")
            day = timestamp.strftime("%Y-%m-%d %A")
            
            # Initialize nested dictionaries if necessary
            if name not in timestamps_by_person:
                timestamps_by_person[name] = {}
            if year not in timestamps_by_person[name]:
                timestamps_by_person[name][year] = {}
            if month not in timestamps_by_person[name][year]:
                timestamps_by_person[name][year][month] = {}
            
            # Initialize timestamps for the day if necessary
            if day not in timestamps_by_person[name][year][month]:
                timestamps_by_person[name][year][month][day] = {'Checkout': None, 'Checkin': None}
            
            # Update Checkin and Checkout timestamps
            if timestamps_by_person[name][year][month][day]['Checkout'] is None or timestamp < datetime.strptime(timestamps_by_person[name][year][month][day]['Checkout'], "%H:%M:%S"):
                timestamps_by_person[name][year][month][day]['Checkout'] = timestamp.strftime("%H:%M:%S")
            if timestamps_by_person[name][year][month][day]['Checkin'] is None or timestamp > datetime.strptime(timestamps_by_person[name][year][month][day]['Checkin'], "%H:%M:%S"):
                timestamps_by_person[name][year][month][day]['Checkin'] = timestamp.strftime("%H:%M:%S")
        
        return timestamps_by_person
    except FileNotFoundError:
        print("File not found. Please provide a valid file path.")
        exit()
    except xlrd.biffh.XLRDError:
        print("Sheet not found. Please provide a valid sheet name.")
        exit()


def export_to_excel(timestamps_by_person, output_file, name=None, year=None, month=None):
    wb = Workbook()
    ws = wb.active

    # Write headers
    headers = ['Name', 'Year', 'Month', 'Day', 'Checkin', 'Checkout']
    for idx, header in enumerate(headers, start=1):
        ws.cell(row=1, column=idx, value=header)

    # Write data
    row_idx = 2
    for person, person_data in timestamps_by_person.items():
        if name is not None and person != name:
            continue
        for person_year, year_data in person_data.items():
            if year is not None and person_year != year:
                continue
            for person_month, month_data in year_data.items():
                if month is not None and person_month != month:
                    continue
                total_hours = 0  # Initialize total hours counter
                for day, timestamps in month_data.items():
                    checkin_time = datetime.strptime(timestamps['Checkin'], "%H:%M:%S")
                    checkout_time = datetime.strptime(timestamps['Checkout'], "%H:%M:%S")
                    total_hours += (checkout_time - checkin_time).total_seconds() / 3600  # Calculate total hours
                    for column_idx, value in enumerate([person, person_year, person_month, day, timestamps['Checkin'], timestamps['Checkout']], start=1):
                        ws.cell(row=row_idx, column=column_idx, value=value)
                    row_idx += 1
                
                # Add a new row for total hours
                ws.cell(row=row_idx, column=1, value='Total Hours')
                ws.cell(row=row_idx, column=6, value=round(total_hours, 2))  # Round total hours to 2 decimal places
                row_idx += 1  # Move to the next row for the next person/month
                
                # Add an extra blank row for readability
                row_idx += 1

    wb.save(f'{output_file}.xlsx')
    print(f'Data exported to {output_file}.xlsx')


def main():
    parser = argparse.ArgumentParser(description='Retrieve timestamps from Excel file and export to Excel.')
    parser.add_argument('--year', '-y', nargs='?', help='Year')
    parser.add_argument('--month', '-m', nargs='?', help='Month')
    parser.add_argument('--output', '-o', default='cicoReport', help='Output file name')

    # Optional argument for person's name
    parser.add_argument('--name', '-n', nargs='?', default=None, help='Name of the person')

    args = parser.parse_args()

    file_path = input("Please provide .xls file path: ") # Update with your file path
    sheet_name = input("Please provide sheet name: ")      # Update with your sheet name
    
    timestamps_by_person = get_timestamps_by_person(file_path, sheet_name)

    export_to_excel(timestamps_by_person, args.output, args.name, args.year, args.month)

if __name__ == "__main__":
    main()
